import os
import openpyxl
from datetime import datetime
from PyQt5.Qt import Qt
from PyQt5.QtWidgets import QDialog, QDialogButtonBox, QVBoxLayout, QGroupBox, QLineEdit

import src.common.settings as config
import src.common.global_vars as gvars
from src.common.functions import adjust_column_length


class ExportParams:
	def __init__(self, model_id):
		self.model_id = model_id

	def interface(self, auto=False):
		# response functions to cancel & ok buttons
		def accept():
			if not auto:
				comment_text = comment_box.text()
			else:
				# TODO: print useful information for multiple fit options
				comment_text = 'placeholder text'

			if config.FILE_LOADED:
				# check if file exists
				input_file_name, input_file_ext = os.path.splitext(config.FULL_FILE_PATH)
				output_file_name = input_file_name + '_params' + input_file_ext
				if os.path.isfile(output_file_name):
					wb = openpyxl.load_workbook(output_file_name)
					if self.model_id == 'cyton1':
						ws = wb['Cyton1 Fit']
					elif self.model_id == 'cyton1.5':
						ws = wb['Cyton1.5 Fit']
					self._export_fit_results(worksheet=ws, comment=comment_text)
					wb.save(output_file_name)
				else:
					wb, c1_ws, c15_ws = self._create_new_workbook()
					if self.model_id == 'cyton1':
						self._export_fit_results(worksheet=c1_ws, comment=comment_text)
					elif self.model_id == 'cyton1.5':
						self._export_fit_results(worksheet=c15_ws, comment=comment_text)
					wb.save(output_file_name)
			else:
				default_file_name = 'cyton_solver_params.xlsx'
				# check for default file at desktop
				default_path = os.path.expanduser('~/Desktop')
				full_default_path = os.path.join(default_path, default_file_name)
				if os.path.isfile(full_default_path):
					wb = openpyxl.load_workbook(full_default_path)
					if self.model_id == 'cyton1':
						ws = wb['Cyton1 Fit']
					elif self.model_id == 'cyton1.5':
						ws = wb['Cyton1.5 Fit']
					self._export_fit_results(worksheet=ws, comment=comment_text)
					wb.save(full_default_path)
				else:
					wb, c1_ws, c15_ws = self._create_new_workbook()
					if self.model_id == 'cyton1':
						self._export_fit_results(worksheet=c1_ws, comment=comment_text)
					elif self.model_id == 'cyton1.5':
						self._export_fit_results(worksheet=c15_ws, comment=comment_text)
					wb.save(full_default_path)
			win.close()

		def reject():
			win.close()

		if not auto:
			win = QDialog()
			_layout = QVBoxLayout()

			# main body of the interface
			gb = QGroupBox('Comment')
			_comment_layout = QVBoxLayout()

			comment_box = QLineEdit()
			comment_box.setFixedWidth(280)
			comment_box.setPlaceholderText(
				"Insert any comment (e.g. free parameter fit)"
			)

			_comment_layout.addWidget(comment_box)
			gb.setLayout(_comment_layout)

			_layout.addWidget(gb)

			# button to finalise the settings
			button = QDialogButtonBox(
				QDialogButtonBox.Ok | QDialogButtonBox.Cancel,
				Qt.Horizontal, win
			)
			button.accepted.connect(accept)
			button.rejected.connect(reject)
			_layout.addWidget(button)

			win.setLayout(_layout)

			win.setAttribute(Qt.WA_DeleteOnClose)
			win.exec_()
		else:
			# automatically accept all incoming save requests for multiple fit options
			accept()

	@staticmethod
	def _create_new_workbook():
		workbook = openpyxl.Workbook()

		# ws_cyton1 = wb.create_sheet(title="Cyton1 Fit")
		worksheet_cyton1 = workbook.active
		worksheet_cyton1.title = "Cyton1 Fit"

		cyton1_headers = [
			'time stamp', 'condition',
			u'\u03bc_0_div', u'\u03c3_0_div',
			u'\u03bc_0_die', u'\u03c3_0_die',
			u'\u03bc_1_div', u'\u03c3_1_div',
			u'\u03bc_1_die', u'\u03c3_1_die',
			'pf0', u'\u03bc_pf', u'\u03c3_pf',
			'Mech. Death Prop.', 'Mech. Death Decay',
			'RSS',
			"Comments"
		]
		worksheet_cyton1.append(cyton1_headers)

		worksheet_cyton15 = workbook.create_sheet(title="Cyton1.5 Fit")
		cyton15_headers = [
			'time stamp', 'condition',
			u'\u03bc_unstim_die', u'\u03c3_unstim_die',
			u'\u03bc_stim_div', u'\u03c3_stim_div',
			u'\u03bc_stim_die', u'\u03c3_stim_die',
			u'\u03bc_stim_dd', u'\u03c3_stim_dd',
			'b', 'pf',
			'RSS',
			'Comments'
		]
		worksheet_cyton15.append(cyton15_headers)

		return workbook, worksheet_cyton1, worksheet_cyton15

	# use existing values stored in gvars
	# so only export selected condition at a time
	def _export_fit_results(self, worksheet, comment):
		now = datetime.now().replace(microsecond=0)
		cond_name = 'Data_free'
		if self.model_id == 'cyton1':
			if config.FILE_LOADED and not config.DATA_FREE:
				cond_name = gvars.CONDITIONS[gvars.C1_ICND]
			contents = [
				now,
				cond_name,
				gvars.C1_PARAMS['mu0div'], gvars.C1_PARAMS['sig0div'],
				gvars.C1_PARAMS['mu0death'], gvars.C1_PARAMS['sig0death'],
				gvars.C1_PARAMS['muSubdiv'], gvars.C1_PARAMS['sigSubdiv'],
				gvars.C1_PARAMS['muSubdeath'], gvars.C1_PARAMS['sigSubdeath'],
				gvars.C1_PARAMS['pf0'], gvars.C1_PARAMS['pfmu'], gvars.C1_PARAMS['pfsig'],
				gvars.C1_PARAMS['MechProp'], gvars.C1_PARAMS['MechDecayConst'],
				gvars.C1_SS,
				comment
			]
		elif self.model_id == 'cyton1.5':
			if config.FILE_LOADED and not config.DATA_FREE:
				cond_name = gvars.CONDITIONS[gvars.C15_ICND]
			contents = [
				now,
				cond_name,
				gvars.C15_PARAMS['unstimMuDeath'], gvars.C15_PARAMS['unstimSigDeath'],
				gvars.C15_PARAMS['stimMuDiv'], gvars.C15_PARAMS['stimSigDiv'],
				gvars.C15_PARAMS['stimMuDeath'], gvars.C15_PARAMS['stimSigDeath'],
				gvars.C15_PARAMS['stimMuDD'], gvars.C15_PARAMS['stimSigDD'],
				gvars.C15_PARAMS['SubDivTime'], gvars.C15_PARAMS['pfrac'],
				gvars.C15_SS,
				comment
			]
		print("[{0}] Saving {1} ({2}) current parameters...".format(now, self.model_id, cond_name), end='')
		content_idx = 0
		max_row = worksheet.max_row
		max_col = worksheet.max_column
		for col in worksheet.iter_cols(min_col=1, min_row=max_row+1, max_col=max_col, max_row=max_row+1):
			for cell in col:
				cell.value = contents[content_idx]
			content_idx += 1
		adjust_column_length(worksheet)
		print("done")
