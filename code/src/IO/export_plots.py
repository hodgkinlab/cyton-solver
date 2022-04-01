import os
import sys
import openpyxl
import numpy as np
import seaborn as sns
import matplotlib.pyplot as plt
from scipy.stats import norm, lognorm, expon
from datetime import datetime

import src.common.settings as config
import src.common.global_vars as gvars
from src.common.functions import adjust_column_length
from src.workbench.cyton1 import Cyton1Model
from src.workbench.cyton15 import Cyton15Model


class ExportPlot:
	def __init__(self, model_id, t0=0, tf=200):
		self.model_id = model_id
		self.path = config.FILE_PATH

		# self.t0 = t0
		# self.tf = tf

	# TODO: there's absolutely no control over plot layouts -> this restriction should be lifted but how should I design it?
	def interface(self):
		pass

	# TODO: complete a set method for plot settings (e.g. last time point, plot size etc)
	def update_settings(self):
		pass

	@staticmethod
	def _compute_pdf(t, mu, sig, lamb=1, pdf_type='Lognormal'):
		if pdf_type == 'Lognormal':
			log_mu = np.log(mu)
			return lognorm.pdf(t, sig, scale=np.exp(log_mu))
		elif pdf_type == 'Gaussian':
			return norm.pdf(t, loc=mu, scale=sig)
		elif pdf_type == 'Exponential':
			return expon.pdf(t, scale=1.0/lamb)

	def _create_folder(self, name):
		self.path = os.path.join(config.FILE_PATH, name)
		if not os.path.exists(self.path):
			os.makedirs(self.path)
		else:
			try:
				# check for existing folder -> put suffix number
				suffix_number = 1
				while os.path.exists(self.path):
					suffix = '{0}_{1:03d}'.format(name, suffix_number)
					suffix_number += 1
					self.path = os.path.join(config.FILE_PATH, suffix)
					# limit number of folders to be 100 per condition
					if suffix_number > 100:
						raise Exception("Number of file copies is too large!")
				os.makedirs(self.path)
			except Exception as e:
				print("[ERROR] {0}".format(e))

	# export an excel file contains all numeric values used in plots
	def _export_plot_table_view(self, t_arr, gens, pdfs, model_results):
		if self.model_id == 'cyton1':
			icnd = gvars.C1_ICND
		elif self.model_id == 'cyton1.5':
			icnd = gvars.C15_ICND

		if config.DATA_FREE:
			icnd = 0

		wb = openpyxl.Workbook()

		"""
		CREATE WORKSHEETS 
			- add headers for simple ones
		"""
		# Probability distribution densities
		ws_pdf = wb.active
		ws_pdf.title = 'PDF'
		ws_pdf_headers = ['time', 'Unstimulated death pdf', 'Stimulated time to first division', 'Stimulated death', 'Stimulated division destiny']
		ws_pdf.append(ws_pdf_headers)  # add headers

		# Total live cells at all time points
		ws_total_live_cells = wb.create_sheet(title='Total Live Cells')
		ws_total_live_cells_headers = ['time', 'live cells']
		ws_total_live_cells.append(ws_total_live_cells_headers)  # add headers

		# Live cells per divisions at harvested time points
		ws_live_cells_per_div_at_ht = wb.create_sheet(title='Live Cells per Divs at ht')

		# Live cells per division at all time points
		# - Total live cells per divisions
		# - Live dividing cells per divisions
		# - Live destiny cells per divisions
		ws_live_cells_per_div = wb.create_sheet(title='Live vs Time all divs')
		ws_live_dividing_cells_per_div = wb.create_sheet(title='Live(dividing) vs Time all divs')
		ws_live_destiny_cells_per_div = wb.create_sheet(title='Live(destiny) vs Time all divs')

		# Total dead cells at all time points
		ws_total_dead_cells = wb.create_sheet(title='Total Dead Cells')
		ws_total_dead_cells_headers = ['time', 'dead cells']
		ws_total_dead_cells.append(ws_total_dead_cells_headers)

		# Dead cells per divisions at harvested time points
		ws_dead_cells_per_div_at_ht = wb.create_sheet(title='Dead Cells per Divs at ht')

		# Dead cells per division at all time points
		#  - Total dead cells
		#  - dead cells from dividng cells
		#  - dead cells from destiny cells
		ws_dead_cells_per_div = wb.create_sheet(title='Dead vs Time all divs')
		ws_dead_dividing_cells_per_div = wb.create_sheet(title='Dead(dividing) vs Time all divs')
		ws_dead_destiny_cells_per_div = wb.create_sheet(title='Dead(destiny) vs Time all divs')

		"""
		Add headers for other worksheets
		"""
		# create headers for "Live/Dead vs Time all divs"
		ws_live_cells_per_div_headers = ['time']
		for gen in gens:
			ws_live_cells_per_div_headers.append('div {0}'.format(gen))
		ws_live_cells_per_div.append(ws_live_cells_per_div_headers)
		ws_live_dividing_cells_per_div.append(ws_live_cells_per_div_headers)
		ws_live_destiny_cells_per_div.append(ws_live_cells_per_div_headers)
		ws_dead_cells_per_div.append(ws_live_cells_per_div_headers)
		ws_dead_dividing_cells_per_div.append(ws_live_cells_per_div_headers)
		ws_dead_destiny_cells_per_div.append(ws_live_cells_per_div_headers)

		# create headers for "Live/Dead Cells per Divs at ht"
		live_dead_per_div_headers = ['divisions']
		for ht in gvars.EXP_HT[icnd]:
			live_dead_per_div_headers.append(ht)
		ws_live_cells_per_div_at_ht.append(live_dead_per_div_headers)
		ws_dead_cells_per_div_at_ht.append(live_dead_per_div_headers)

		"""
		Populate contents
		"""
		for itpt, t in enumerate(t_arr):
			pdf = [t, pdfs[0][itpt], pdfs[1][itpt], pdfs[2][itpt], pdfs[3][itpt]]
			ws_pdf.append(pdf)

			live = [t, model_results[1][0][itpt]]
			ws_total_live_cells.append(live)

			live_per_div = [t]
			live_dividing_per_div = [t]
			live_destiny_per_div = [t]

			dead = [t, model_results[2][0][itpt]]
			ws_total_dead_cells.append(dead)

			dead_per_div = [t]
			dead_dividing_per_div = [t]
			dead_destiny_per_div = [t]

			for igen in range(gvars.MAX_DIV_PER_CONDITIONS[icnd]+1):
				live_per_div.append(model_results[1][1][igen][itpt])
				live_dividing_per_div.append(model_results[1][2][igen][itpt])
				live_destiny_per_div.append(model_results[1][3][igen][itpt])

				dead_per_div.append(model_results[2][1][igen][itpt])
				dead_dividing_per_div.append(model_results[2][2][igen][itpt])
				dead_destiny_per_div.append(model_results[2][3][igen][itpt])

			ws_live_cells_per_div.append(live_per_div)
			ws_live_dividing_cells_per_div.append(live_dividing_per_div)
			ws_live_destiny_cells_per_div.append(live_destiny_per_div)

			ws_dead_cells_per_div.append(dead_per_div)
			ws_dead_dividing_cells_per_div.append(dead_dividing_per_div)
			ws_dead_destiny_cells_per_div.append(dead_destiny_per_div)

		"""
		(optional) adjust all columns to content size if possible
		"""
		adjust_column_length(ws_pdf)
		adjust_column_length(ws_total_live_cells)
		adjust_column_length(ws_live_cells_per_div)
		adjust_column_length(ws_live_dividing_cells_per_div)
		adjust_column_length(ws_live_destiny_cells_per_div)
		adjust_column_length(ws_dead_cells_per_div)
		adjust_column_length(ws_dead_dividing_cells_per_div)
		adjust_column_length(ws_dead_destiny_cells_per_div)

		for igen in range(gvars.MAX_DIV_PER_CONDITIONS[icnd]+1):
			live_tmp = ['div {0}'.format(igen)]
			dead_tmp = ['div {0}'.format(igen)]
			for itpt in range(len(gvars.EXP_HT[icnd])):
				live_tmp.append(model_results[0][1][itpt][igen])
				dead_tmp.append(model_results[0][2][itpt][igen])
			ws_live_cells_per_div_at_ht.append(live_tmp)
			ws_dead_cells_per_div_at_ht.append(dead_tmp)
		adjust_column_length(ws_live_cells_per_div_at_ht)
		adjust_column_length(ws_dead_cells_per_div_at_ht)

		wb.save(self.path + '/plot_table_view.xlsx')

	# PROTOTYPE PLOT EXPORTING FUNCTION - no control but pre-designed plots only
	def export_plot_with_data(self):
		now = datetime.now().replace(microsecond=0)

		# set seaborn plot settings
		sns.set(context='paper')
		sns.set_style('whitegrid')

		try:
			if self.model_id == 'cyton1':
				# TODO: Complete cyton 1 path
				# icnd = gvars.C1_ICND
				# cp = sns.hls_palette(gvars.MAX_DIV_PER_CONDITIONS[icnd], l=0.4, s=0.5)
				raise NotImplementedError("Exporting plot for Cyton 1 is not implemented yet!")
			elif self.model_id == 'cyton1.5':
				if config.DATA_FREE:
					icnd = 0
					cond_name = 'Data Free'
					file_name = 'Data Free'
				else:
					icnd = gvars.C15_ICND
					cond_name = gvars.CONDITIONS[icnd]
					file_name = gvars.CONDITIONS[icnd]
				# create seaborn HLS color palette
				cp = sns.hls_palette(gvars.MAX_DIV_PER_CONDITIONS[icnd]+1, l=0.4, s=0.5)

				# check user os
				platform = sys.platform
				if '/' in cond_name and platform == 'win32':
					# need to change this for Windows environment : reserved characters for file names are <, >, :, ", /, \, |, ?, *
					file_name = file_name.replace('/', '_')
				elif '/' in cond_name and platform == 'darwin':
					# mac automatically converts : to / in file name
					file_name = file_name.replace('/', ':')
				self._create_folder(file_name)

				# set default time frame
				t0 = 0.0
				# tf = max(gvars.EXP_HT[icnd])
				tf = max(gvars.HT)

				# create (x, y) arrays
				time_arr = np.linspace(t0, tf, num=int(tf/gvars.TIME_INC)+1)
				gens = [i for i in range(gvars.MAX_DIV_PER_CONDITIONS[icnd]+1)]

				# compute Cyton model
				model_results = Cyton15Model(
					gvars.EXP_HT[gvars.C15_ICND],
					gvars.INIT_CELL,
					gvars.MAX_DIV_PER_CONDITIONS[icnd],
					gvars.TIME_INC, [], [], False
				).compute_model_results(time_arr, gvars.C15_PARAMS)

				# sort results obtained above
				total_live_cells = model_results[1][0]
				live_cells_gen = model_results[1][1]
				live_cells_gen_at_ht = model_results[0][1]  # summary of model curve at harvested times
				live_dividing_cells = model_results[1][2]
				live_destiny_cells = model_results[1][3]
				live_unstimulated_cells = model_results[1][4]

				# compute for cumulative live cells -> easy to visualise transition from dividing -> destiny
				cumulative_dividing = np.zeros(shape=time_arr.size)
				cumulative_destiny = np.zeros(shape=time_arr.size)
				for j in range(time_arr.size):
					for igen in range(gvars.MAX_DIV_PER_CONDITIONS[icnd]+1):
						cumulative_dividing[j] += live_dividing_cells[igen, j]
						cumulative_destiny[j] += live_destiny_cells[igen, j]

				total_dead_cells = model_results[2][0]
				dead_cells_gen = model_results[2][1]
				dead_dividing_cells = model_results[2][2]
				dead_destiny_cells = model_results[2][3]
				dead_unstimulated_cells = model_results[2][4]

				params = gvars.C15_PARAMS
				# pdf types for parameters
				unstim_die_pdf = config.CYTON15_CONFIG['unstim_die']
				stim_div_pdf = config.CYTON15_CONFIG['stim_div']
				stim_die_pdf = config.CYTON15_CONFIG['stim_die']
				stim_dd_pdf = config.CYTON15_CONFIG['stim_dd']

				# compute PDFs
				unstim_die = self._compute_pdf(time_arr, params['unstimMuDeath'], params['unstimSigDeath'], pdf_type=unstim_die_pdf)
				stim_tfd = self._compute_pdf(time_arr, params['stimMuDiv'], params['stimSigDiv'], pdf_type=stim_div_pdf)
				stim_die = self._compute_pdf(time_arr, params['stimMuDeath'], params['stimSigDeath'], pdf_type=stim_die_pdf)
				stim_des = self._compute_pdf(time_arr, params['stimMuDD'], params['stimSigDD'], pdf_type=stim_dd_pdf)

				"""PLOT TYPE I : COMBINE TOTAL CELLS & PDFs"""
				fig, (pdf, cell_dynamic) = plt.subplots(1, 2, figsize=(16, 8))
				fig.suptitle('{0} : {1}'.format(config.FILE_NAME, cond_name))
				pdf.set_title('Probability Distributions')
				pdf.set_xlabel('time (hrs)')
				pdf.set_ylabel('density')
				# pdf.set_xlim(min(ht), max(ht))

				pdf.fill_between(time_arr, -unstim_die, color='orange', label='Death of unstimulated cells',alpha=0.5)
				pdf.fill_between(time_arr, stim_tfd, color='blue', label='Time to first division', alpha=0.5)
				pdf.fill_between(time_arr, -stim_die, color='red', label='Death of stimulated cells', alpha=0.5)
				pdf.fill_between(time_arr, stim_des, color='green', label='Division destiny', alpha=0.5)
				pdf.legend()

				cell_dynamic.set_title('Total Live Cells')
				cell_dynamic.set_xlabel('time (hrs)')
				cell_dynamic.set_ylabel('number of cells')
				# tot.set_xlim(min(ht), max(ht))
				# plot live cell : total & cells at all generations

				# do not include data points for data free mode
				if not config.DATA_FREE:
					# plot data points with SEM error bar
					(datapoints, caps, _) = cell_dynamic.errorbar(
						gvars.EXP_HT[icnd], gvars.TOTAL_CELLS[icnd],
						yerr=gvars.TOTAL_CELLS_SEM[icnd],
						color=(.6, 0, 0, 1.0),
						fmt='o--',
						linewidth=1.0,
						markersize=5,
						capsize=4,
						label='Data : %s' % cond_name
					)

					for cap in caps:
						cap.set_markeredgewidth(1)

				cell_dynamic.plot(time_arr, total_live_cells, '-', c='blue', linewidth=1.5, label='total live cells')
				cell_dynamic.plot(time_arr, live_unstimulated_cells, '--', c='orange', label='unstimulated cells')
				for igen in range(gvars.MAX_DIV_PER_CONDITIONS[icnd]+1):
					cell_dynamic.fill_between(time_arr, live_cells_gen[igen], color=cp[igen], alpha=0.5, label='gen {0}'.format(igen))
				cell_dynamic.plot(time_arr, cumulative_dividing, 'b:', label='dividing cells')
				cell_dynamic.plot(time_arr, cumulative_destiny, 'g:', label='dsetiny cells')

				# organise legend items
				if not config.DATA_FREE:
					handles, labels = cell_dynamic.get_legend_handles_labels()
					line_handles = [handles[len(handles)-1], handles[0], handles[1], handles[2], handles[3]]
					line_labels = [labels[len(labels)-1], labels[0], labels[1], labels[2], labels[3]]
					block_handles = handles[4:4+gvars.MAX_DIV_PER_CONDITIONS[icnd]+1]
					block_labels = labels[4:4+gvars.MAX_DIV_PER_CONDITIONS[icnd]+1]
				else:
					handles, labels = cell_dynamic.get_legend_handles_labels()
					line_handles = [handles[0], handles[1], handles[2], handles[3]]
					line_labels = [labels[0], labels[1], labels[2], labels[3]]
					block_handles = handles[4:4+gvars.MAX_DIV_PER_CONDITIONS[icnd]+1]
					block_labels = labels[4:4+gvars.MAX_DIV_PER_CONDITIONS[icnd]+1]

				plt.gca().add_artist(cell_dynamic.legend(line_handles, line_labels, loc=1))
				cell_dynamic.legend(block_handles, block_labels, loc=2)

				plt.savefig(self.path + '/{0}_{1}.pdf'.format('fig1', file_name))

				"""PLOT TYPE II : CELLS PER GENERATION AT HARVESTED TIME"""
				fig2 = plt.figure(figsize=(16, 8))
				fig2.suptitle('{0} : {1}'.format(config.FILE_NAME, cond_name))
				fig2.text(0.5, 0.04, "Generations", ha='center', va='center')
				fig2.text(0.05, 0.5, "Cell Number", ha='center', va='center', rotation=90)

				# increase number of columns in case of large time point data
				if gvars.EXP_NUM_TP[icnd] > 12:
					num_rows = 3
					num_cols = 7
				else:
					num_rows = np.ceil(len(gvars.EXP_HT[icnd])/3)
					num_cols = 3

				for itpt, t in enumerate(gvars.EXP_HT[icnd]):
					ax = plt.subplot(num_rows, num_cols, itpt+1)
					ax.set_title('time at {0} hrs'.format(t))
					ax.plot(gens, live_cells_gen_at_ht[itpt], 'o-', c='blue', label='fitted model')

					if not config.DATA_FREE:
						(datapoints, caps, _) = ax.errorbar(
							gens, gvars.CELL_GENS[icnd][itpt],
							yerr=gvars.CELL_GENS_SEM[icnd][itpt],
							color=(.6, 0, 0, 1.0),
							fmt='o--',
							linewidth=1.0,
							markersize=5,
							capsize=4,
							label='Data : %s' % cond_name
						)

						for cap in caps:
							cap.set_markeredgewidth(1)

				if not config.DATA_FREE:
					handles, labels = ax.get_legend_handles_labels()
					data_line_handles = [handles[1], handles[0]]
					data_line_labels = [labels[1], labels[0]]
				else:
					handles, labels = ax.get_legend_handles_labels()
					data_line_handles = [handles[0]]
					data_line_labels = [labels[0]]
				fig2.legend(data_line_handles, data_line_labels)
				fig2.tight_layout(rect=(0.05, 0.05, 0.95, 0.95))

				plt.savefig(self.path + '/{0}_{1}.pdf'.format('fig2', file_name))

				"""PLOT TYPE III : MODEL DETAILS (DEATH, DIVIDING, DESINTY)"""
				fig3, [(ax1, ax2), (ax3, ax4)] = plt.subplots(2, 2, figsize=(18, 10))

				# fig.tight_layout()
				fig3.suptitle('{0} : {1}'.format(config.FILE_NAME, cond_name))
				ax1.set_title('Number of Dividing cells')
				ax1.set_xlabel('time (hrs)')
				ax1.set_ylabel('number of cells')

				ax2.set_title('Number of Destiny cells')
				ax2.set_xlabel('time (hrs)')
				ax2.set_ylabel('number of cells')

				ax3.set_title('Number of Dead cells')
				ax3.set_xlabel('time (hrs)')
				ax3.set_ylabel('number of cells')

				ax4.set_title('Total Dead Cells')
				ax4.set_xlabel('time (hrs)')
				ax4.set_ylabel('number of cells')

				ax3.plot([], [], '-', color='k', label='dead diving cells')
				ax3.plot([], [], '--', color='k', label='dead destiny cells')
				for igen in range(gvars.MAX_DIV_PER_CONDITIONS[icnd]+1):
					ax1.fill_between(time_arr, live_dividing_cells[igen], color=cp[igen], alpha=0.5, label='gen {0}'.format(igen))
					ax2.fill_between(time_arr, live_destiny_cells[igen], color=cp[igen], alpha=0.5, label='gen {0}'.format(igen))
					ax3.fill_between(time_arr, dead_cells_gen[igen], color=cp[igen], alpha=0.5, label='gen {0}'.format(igen))
					ax3.plot(time_arr, dead_dividing_cells[igen], '-', color=cp[igen])
					ax3.plot(time_arr, dead_destiny_cells[igen], '--', color=cp[igen])
					ax4.fill_between(time_arr, dead_cells_gen[igen], color=cp[igen], alpha=0.5, label='gen {0}'.format(igen))
				ax4.plot(time_arr, total_dead_cells, 'r-', linewidth=1.5, label='total dead cells')
				ax4.plot(time_arr, dead_unstimulated_cells, 'r--', label='dead unstimulated cells')
				ax1.legend()
				ax2.legend()
				ax3.legend()
				ax4.legend()

				plt.savefig(self.path + '/{0}_{1}.pdf'.format('fig3', file_name))

				"""PLOT TYPE IV : TOTAL LIVE CELL PER GENERATION OVER TIME"""
				fig4 = plt.figure(figsize=(16, 8))
				fig4.suptitle('{0} : {1}'.format(config.FILE_NAME, cond_name))
				fig4.text(0.5, 0.04, "time (hrs)", ha='center', va='center')
				fig4.text(0.05, 0.5, "number of live cell", ha='center', va='center', rotation=90)
				num_rows = np.ceil(int(gvars.MAX_DIV_PER_CONDITIONS[icnd]+1) / 3)
				for igen in range(gvars.MAX_DIV_PER_CONDITIONS[icnd]+1):
					ax = plt.subplot(num_rows, 3, igen + 1)
					ax.set_title('gen {0}'.format(igen))
					ax.fill_between(time_arr, live_cells_gen[igen], color=cp[igen], alpha=0.5)
					if not config.DATA_FREE:
						tmp_data = [sublist[igen] for sublist in gvars.CELL_GENS[icnd]]
						tmp_err = [sublist[igen] for sublist in gvars.CELL_GENS_SEM[icnd]]
						# ax.plot(GV.harvested_time_default, tmp_data, 'o--', color=color_palette[igen])
						(datapoints, caps, _) = ax.errorbar(
							gvars.EXP_HT[icnd], tmp_data,
							yerr=tmp_err,
							color=cp[igen],
							fmt='o--',
							linewidth=1.0,
							markersize=5,
							capsize=4,
							label='Data : %s' % cond_name
						)
						for cap in caps:
							cap.set_markeredgewidth(1)

				fig4.tight_layout(rect=(0.05, 0.05, 0.95, 0.95))
				plt.savefig(self.path + '/{0}_{1}.pdf'.format('fig4-1', file_name))

				# same as above but in a single panel
				fig5 = plt.figure(figsize=(16, 8))
				fig5.suptitle('{0} : {1}'.format(config.FILE_NAME, cond_name))
				for igen in range(gvars.MAX_DIV+1):
					tmp_data = [sublist[igen] for sublist in gvars.CELL_GENS[icnd]]
					tmp_err = [sublist[igen] for sublist in gvars.CELL_GENS_SEM[icnd]]
					(datapoints, caps, _) = plt.errorbar(
						gvars.HT, tmp_data,
						yerr=tmp_err,
						color=cp[igen],
						fmt='o--',
						linewidth=1.0,
						markersize=5,
						capsize=4,
						# label='Data : %s' % cond_name
					)
					for cap in caps:
						cap.set_markeredgewidth(1)
					plt.fill_between(time_arr, live_cells_gen[igen], color=cp[igen], alpha=0.4, label='gen {0}'.format(igen))
				plt.xlabel('time (hrs)')
				plt.ylabel('number of cells')
				plt.legend()
				plt.savefig(self.path + '/{0}_{1}.pdf'.format('fig4-2', file_name))

				print('[{0}] Plots exported in "{1}"'.format(now, self.path))

				# export plot in table views
				self._export_plot_table_view(time_arr, gens, [unstim_die, stim_tfd, stim_die, stim_des], model_results)

		except NotImplementedError as nie:
			print("[NOT IMPLEMENTED ERROR] {0}".format(nie))
		except Exception as e:
			print("[ERROR] {0}".format(e))

