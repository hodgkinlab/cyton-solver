import openpyxl
from datetime import datetime
from itertools import groupby

from src.IO.error_handler import CustomError
from src.common.functions import remove_empty


class ReadData:
	def __init__(self, file):
		# define workbook from input excel file
		#  - "data_only" parameter is necessary to remove equations
		workbook = openpyxl.load_workbook(file, data_only=True)

		# define specific worksheet from workbook
		#  - this potentially can be extended for multiple worksheets under one excel file
		#  - TODO: perhaps I could implement an interface for a user to choose worksheet
		sheet_names = workbook.sheetnames

		try:
			# try Legacy data format by Cohort Explorer v2018.05.09
			ws = workbook['Input for CytonSolver']
		except Exception as e:
			# otherwise default is first sheet of FormatHLSep2016
			ws = workbook[sheet_names[0]]  # only pick first worksheet for now

		# define maximum row / col numbers in the excel file. For some reason this is a bit buggy
		# self.max_row = ws.max_row
		# self.max_col = ws.max_column

		# validate input data file : boolean type
		#  - 1. FormatHLSep2016 : True
		#  - 2. Legacy : False
		data_format = self.validate_data_format(ws)

		# collect meta information of the experiment
		self.meta_information = self.get_experiment_info(ws, data_format)
		self.condition_names = self.get_condition_names(ws, data_format)
		self.generation_per_condition = self.get_generation_information(ws, data_format)
		self.harvested_times, self.harvested_times_reps, self.num_time_points = self.get_time_points(ws, data_format)
		self.data = self.get_cell_number(ws, data_format)

	# def __del__(self):
	#     print("Deleting ReadData object")

	# NOTE: this would only work if people are agree on the format itself
	@staticmethod
	def validate_data_format(worksheet):
		"""
		validate input excel file data format

		:param1 worksheet: (Worksheet) openpyxl worksheet object
		:returns: True if 'FormatHLSep2016', False otherwise
		"""
		format_tag = worksheet['B1']
		if format_tag.value == 'HL-SEP-2016':
			return True
		else:
			return False

	@staticmethod
	def get_experiment_info(worksheet, data_format):
		"""
		collects user defined experiment information.
		This only provides a guide to build the data scheme, however, crucial to have it in the first place.

		:param1 worksheet: (Worksheet) openpyxl worksheet object
		:param2 data_format: (bool) validation of data format
		:returns: (dict) experiments basic information
		:raises CustomError: raises a CustomError (inherits Exception) for missing critical information
		"""
		if data_format:
			# initialise information dictionary
			exp_info = {
				'num_tp': 0,
				'last_gen': 0,
				'num_condition': 0,
				'num_replicate': 0,
				'num_beads': 0,
				'prop_bead_gated': 0
			}
			col_idx = 0
			# iterate through rows [B3 - B8] where experiment informations are located.
			# col_idx is essentially a check index
			for col_cell_object in worksheet.iter_rows(min_col=2, max_col=2, min_row=3, max_row=8):
				for cell in col_cell_object:
					val = cell.value
					# if cells are empty raise an Exception. This is to guide user to fix their data.
					if val is None:
						if col_idx == 0:
							raise CustomError("Number of time points not set")
						elif col_idx == 1:
							raise CustomError("Last generation not set")
						elif col_idx == 2:
							raise CustomError("Number of conditions not set")
						elif col_idx == 3:
							raise CustomError("Number of replicates not set")
						elif col_idx == 4:
							raise CustomError("Number of beads not set")
						elif col_idx == 5:
							raise CustomError("Prop. beads gated not set")
					else:
						# validate if data make sense
						if col_idx == 0:
							if val <= 0:
								raise CustomError("Number of time points cannot be 0 or negative")
							else:
								exp_info['num_tp'] = val
						elif col_idx == 1:
							if val <= 0:
								raise CustomError("Last generation must be greater than 0")
							else:
								exp_info['last_gen'] = val
						elif col_idx == 2:
							if val <= 0:
								raise CustomError("Number of condition must to be greater than 0")
							else:
								exp_info['num_condition'] = val
						elif col_idx == 3:
							if val <= 0:
								raise CustomError("Number of replicate must be greater than 0")
							else:
								exp_info['num_replicate'] = val
						elif col_idx == 4:
							if val <= 0:
								raise CustomError("Number of beads must be greater than 0")
							else:
								exp_info['num_beads'] = val
						elif col_idx == 5:
							if 0.0 <= val <= 1.0:
								exp_info['prop_bead_gated'] = val
							else:
								raise CustomError("Prop. beads gated must be in range 0 and 1")
					col_idx += 1
			return exp_info
		else:
			# initialise information dictionary for legacy type data format
			exp_info = {
				'stimulus': '',
				'date': '',
				'cell_type': '',
				'comment': ''
			}
			# iterate through first row [A1 - max_col1]
			col_idx = 0
			for row_cell_object in worksheet.iter_rows(min_col=1, max_col=worksheet.max_column, min_row=1, max_row=1):
				for cell in row_cell_object:
					val = cell.value
					if val is not None:
						if val == 'Stimulus' or val == 'stimulus':
							exp_info['stimulus'] = row_cell_object[col_idx + 1].value
						elif val == 'Date' or val == 'date':
							# if date is not specified default to current time
							if row_cell_object[col_idx + 1].value is None:
								ts = str(datetime.now().replace(microsecond=0))
								exp_info['date'] = ts
							else:
								exp_info['date'] = row_cell_object[col_idx + 1].value
						elif val == 'Cell' or val == 'cell':
							exp_info['cell_type'] = row_cell_object[col_idx + 1].value
						elif val == 'Comment' or val == 'comment':
							exp_info['comment'] = row_cell_object[col_idx + 1].value
					else:
						if val is None:
							pass
						elif val is not None and col_idx == 0:
							raise CustomError("Missing stimulus name!")
						elif val is not None and col_idx == 1:
							print("Missing experiment date time. Recording it to current time...")
							ts = str(datetime.now().replace(microsecond=0))
							exp_info['date'] = ts
						elif val is not None and col_idx == 2:
							raise CustomError("Missing cell type!")
						elif val is not None and col_idx == 3:
							comment = 'No comment'
							exp_info['comment'] = comment
					col_idx += 1
			return exp_info

	@staticmethod
	def get_condition_names(worksheet, data_format):
		"""
		collects experiment condition names. (e.g. Wildtype-1U, Wildtype-Unstim)

		:param1 worksheet: (Worksheet) openpyxl worksheet object
		:param2 data_format: (bool) validation of data format
		:returns: (list) condition names
		"""
		conditions = []
		if data_format:
			cells = worksheet['A12':'A' + str(worksheet.max_row)]
			for cell in cells:
				val = cell[0].value
				if val is not None:
					# conver to string in case of rare numeric condition name
					conditions.append(str(val))
				else:
					break
			return conditions
		else:
			for col_cell_object in worksheet.iter_cols(min_col=1, max_col=1, min_row=2, max_row=worksheet.max_row):
				for cell in col_cell_object:
					val = cell.value
					if val is not None:
						conditions.append(str(val))
			return conditions

	def get_generation_information(self, worksheet, data_format):
		"""
		collects last generation information per conditions.

		:param1 worksheet: (Worksheet) openpyxl worksheet object
		:param2 data_format: (bool) validation of data format
		:returns: (list) last generation per condition indexed by exact order of condition list
		"""
		generation = []

		if data_format:
			# there's no way to identify last generation per condition in new data format
			# just simply assume that all conditions are sharing same last generation that user set
			for idx in range(len(self.condition_names)):
				generation.append(self.meta_information['last_gen'])
			return generation
		else:
			counter = 0
			for col_cell_object in worksheet.iter_cols(min_col=2, max_col=2, min_row=2, max_row=worksheet.max_row + 1):
				for cell in col_cell_object:
					val = cell.value
					if val is None:
						generation.append(counter - 2)
						counter = 0
					else:
						counter += 1
			return generation

	def get_time_points(self, worksheet, data_format):
		"""
		collects harvested time points per conditions.

		:param1 worksheet: (Worksheet) openpyxl worksheet object
		:param2 data_format: (bool) validation of data format
		:returns: (tuple; list, list) returns harvested time points & duplicates of harvested time points to represent replicates
		"""

		# initialise with empty list with same dimension as get_condition_names. This allows to access time point data sorted by condition index.
		harvested_times_reps = [
			[] for _ in range(len(self.condition_names))
		]
		if data_format:
			cells = worksheet['D2':'D' + str(worksheet.max_row)]
			# brings sample_names column from the data to check
			condition_column = worksheet['E2':'E' + str(worksheet.max_row)]
			for i, cell in enumerate(cells):
				val = cell[0].value
				curr_condition_name = str(condition_column[i][0].value)
				# if cell has a value means time to move onto next time point
				if val is not None:
					tmp = val
					if curr_condition_name in self.condition_names:
						target_condition_idx = self.condition_names.index(curr_condition_name)
						harvested_times_reps[target_condition_idx].append(val)
				else:
					if curr_condition_name in self.condition_names:
						target_condition_idx = self.condition_names.index(curr_condition_name)
						harvested_times_reps[target_condition_idx].append(tmp)
			harvested_times = [
				[] for _ in range(len(self.condition_names))
			]
			num_time_points = [
				[] for _ in range(len(self.condition_names))
			]
			# remove duplicates in harvested_time_reps
			for i in range(len(self.condition_names)):
				harvested_times[i] = set(harvested_times_reps[i])
				harvested_times[i] = list(sorted(harvested_times[i]))
				num_time_points[i] = len(harvested_times[i])

			return harvested_times, harvested_times_reps, num_time_points
		else:
			harvested_times_reps = [
				[] for _ in range(len(self.condition_names))
			]

			icnd = 0
			itpt = 2
			next_condition_break = self.generation_per_condition[icnd] + 3
			for row_cell_object in worksheet.iter_rows(min_col=2, max_col=worksheet.max_column, min_row=2, max_row=worksheet.max_row):
				for cell in row_cell_object:
					val = cell.value
					if (cell.row == itpt) and val is not None:
						harvested_times_reps[icnd].append(val)

				# a condition check for moving onto next condition & update icnd, itpt
				if cell.row % next_condition_break == 0:
					itpt += self.generation_per_condition[icnd] + 3
					icnd += 1
					if icnd > len(self.condition_names) - 1:
						break
					next_condition_break += self.generation_per_condition[icnd] + 3

			harvested_times = [
				[] for _ in range(len(self.condition_names))
			]
			num_time_points = [
				[] for _ in range(len(self.condition_names))
			]
			for i in range(len(self.condition_names)):
				harvested_times[i] = set(harvested_times_reps[i])
				harvested_times[i] = list(sorted(harvested_times[i]))
				num_time_points[i] = len(harvested_times[i])

			return harvested_times, harvested_times_reps, num_time_points

	def get_cell_number(self, worksheet, data_format):
		"""
		collects main data of the experiment.
		Sorts cell number data according to icnd, itpt, igen

		:param1 worksheet: (Worksheet) openpyxl worksheet object
		:param2 data_format: (bool) validation of data format
		:returns: (list) cell number information
		:raises ArithmeticError: it's possible to encounter divide by 0 case
		"""

		if data_format:
			dataset = [
				[
					[
						[] for _ in range(self.meta_information['last_gen'] + 1)
					] for _ in range(self.meta_information['num_tp'])
				] for _ in range(len(self.condition_names))
			]

			exp_beads = self.meta_information['num_beads'] * self.meta_information['prop_bead_gated']

			itpt = 0
			row_counter = 0

			time_list = worksheet['D2':'D' + str(worksheet.max_row)]
			condition_list = worksheet['E2':'E' + str(worksheet.max_row)]
			bead_list = worksheet['F2':'F' + str(worksheet.max_row)]
			for row_cell_object in worksheet.iter_rows(min_col=8, max_col=worksheet.max_column, min_row=2, max_row=worksheet.max_row):
				for col_idx, cell in enumerate(row_cell_object):
					val = cell.value
					curr_condition_name = str(condition_list[row_counter][0].value)
					curr_time = time_list[row_counter][0].value
					curr_beads = bead_list[row_counter][0].value
					# time point update criteria
					if curr_time is not None and row_counter > 0 and col_idx == 0:
						itpt += 1
					# iteratively check if condition name exists in self.condition_name.
					# This way user can manually take out one of replicates without deleteing entire row, and system will dynamically adjust matrix size
					if curr_condition_name in self.condition_names:
						target_condition_idx = self.condition_names.index(curr_condition_name)
						if curr_beads is None:
							raise ArithmeticError(
								"Empty bead number is detected. Check your number of beads at : time point %.2f, condition %s" % (
									self.harvested_times[target_condition_idx][itpt], curr_condition_name))
						elif float(curr_beads) == 0:
							raise ArithmeticError(
								"Divide by zero issue detected. Check your number of beads at : time point %.2f, condition %s" % (
									self.harvested_times[target_condition_idx][itpt], curr_condition_name))
						if val is not None:
							val = val * exp_beads / curr_beads
						else:
							val = 0
						dataset[target_condition_idx][itpt][col_idx].append(val)

					if col_idx == self.meta_information['last_gen']:
						row_counter += 1

			return remove_empty(dataset)
		else:
			# define a mirrored list of harvest_time_reps to counts for number of replicates per condition per time point
			NUMBER_OF_REPLICATES = []
			for icnd in range(len(self.condition_names)):
				NUMBER_OF_REPLICATES.append(
					[len(list(group)) for key, group in groupby(self.harvested_times_reps[icnd])])

			dataset = [
				[
					[
						[] for _ in range(max(self.generation_per_condition) + 1)
					] for _ in range(max(self.num_time_points))
				] for _ in range(len(self.condition_names))
			]

			icnd, itpt, igen = 0, 0, 0
			replicate_counter = 0
			time_row = 2
			next_condition_break = self.generation_per_condition[icnd] + 4
			for row_cell_object in worksheet.iter_rows(min_col=2, max_col=worksheet.max_column, min_row=2, max_row=worksheet.max_row):
				for cell in row_cell_object:
					val = cell.value
					if cell.row != time_row and cell.row != next_condition_break:
						if val is not None:
							dataset[icnd][itpt][igen].append(val)
							replicate_counter += 1
							if cell.col_idx == (len(self.harvested_times_reps[icnd]) + 1):
								igen += 1
							if replicate_counter % NUMBER_OF_REPLICATES[icnd][itpt] == 0:
								itpt += 1
								replicate_counter = 0
					if cell.col_idx == (len(self.harvested_times_reps[icnd]) + 1):
						itpt = 0
						if cell.row % next_condition_break == 0:
							time_row += self.generation_per_condition[icnd] + 3
							icnd += 1
							igen = 0
							if icnd > len(self.condition_names) - 1:
								break
							next_condition_break += self.generation_per_condition[icnd] + 3

			return remove_empty(dataset)
