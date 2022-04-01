import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils.exceptions import InvalidFileException, SheetTitleException
from datetime import datetime
from PyQt5.Qt import Qt
from PyQt5.QtWidgets import QFileDialog, QDialog, QDialogButtonBox, QTableWidget, QGridLayout, QComboBox, QTableWidgetItem, QAbstractItemView, QAbstractScrollArea, QHeaderView

import src.common.settings as config
import src.common.global_vars as gvars
from src.IO.error_handler import CustomError


class ImportParams:
	def __init__(self, model_id, parent=None):
		self.model_id = model_id
		self.parent = parent
		self.workbook = None
		self.worksheet = None

	def interface(self):
		# validate imported data & param excel files are of same experiments
		# def validate_file():
		# 	condition_names = []
		# 	for item in self.worksheet.iter_rows(min_col=2, max_col=2, min_row=2, max_row=self.worksheet.max_row):
		# 		for cell in item:
		# 			print(cell.value)
		# 	if 0:
		# 		return False
		# 	return True

		def try_default_path():
			file_name, file_ext = os.path.splitext(config.FULL_FILE_PATH)
			f = file_name + '_params' + file_ext

			if os.path.isfile(f):
				self.workbook = openpyxl.load_workbook(f)
				if self.model_id == 'cyton1':
					self.worksheet = self.workbook['Cyton1 Fit']
				elif self.model_id == 'cyton1.5':
					self.worksheet = self.workbook['Cyton1.5 Fit']
				return True
			else:
				return False

		def accept():
			if self.model_id == 'cyton1':
				data_to_update = data[l.currentIndex()][1:14]
				gvars.C1_PARAMS['mu0div'] = data_to_update[0]
				gvars.C1_PARAMS['sig0div'] = data_to_update[1]
				gvars.C1_PARAMS['mu0death'] = data_to_update[2]
				gvars.C1_PARAMS['sig0death'] = data_to_update[3]
				gvars.C1_PARAMS['muSubdiv'] = data_to_update[4]
				gvars.C1_PARAMS['sigSubdiv'] = data_to_update[5]
				gvars.C1_PARAMS['muSubdeath'] = data_to_update[6]
				gvars.C1_PARAMS['sigSubdeath'] = data_to_update[7]
				gvars.C1_PARAMS['pf0'] = data_to_update[8]
				gvars.C1_PARAMS['pfmu'] = data_to_update[9]
				gvars.C1_PARAMS['pfsig'] = data_to_update[10]
				gvars.C1_PARAMS['MechProp'] = data_to_update[11]
				gvars.C1_PARAMS['MechDecayConst'] = data_to_update[12]

				for i, item in enumerate(self.parent.params):
					item.valueChanged.disconnect(self.parent.c1_event_handler.handle_param_value_changed)
					item.setValue(data_to_update[i])
					item.valueChanged.connect(self.parent.c1_event_handler.handle_param_value_changed)
				self.parent.c1_plot.update_plot()
			elif self.model_id == 'cyton1.5':
				data_to_update = data[l.currentIndex()][1:11]

				gvars.C15_PARAMS['unstimMuDeath'] = data_to_update[0]
				gvars.C15_PARAMS['unstimSigDeath'] = data_to_update[1]
				gvars.C15_PARAMS['stimMuDiv'] = data_to_update[2]
				gvars.C15_PARAMS['stimSigDiv'] = data_to_update[3]
				gvars.C15_PARAMS['stimMuDeath'] = data_to_update[4]
				gvars.C15_PARAMS['stimSigDeath'] = data_to_update[5]
				gvars.C15_PARAMS['stimMuDD'] = data_to_update[6]
				gvars.C15_PARAMS['stimSigDD'] = data_to_update[7]
				gvars.C15_PARAMS['SubDivTime'] = data_to_update[8]
				gvars.C15_PARAMS['pfrac'] = data_to_update[9]

				for i, item in enumerate(self.parent.params15):
					item.valueChanged.disconnect(self.parent.c15_event_handler.handle_param_value_changed)
					item.setValue(data_to_update[i])
					item.valueChanged.connect(self.parent.c15_event_handler.handle_param_value_changed)
				self.parent.c15_plot.update_plot()
			win.close()

		def reject():
			win.close()

		def handle_data_list():
			for i, datum in enumerate(data[l.currentIndex()]):
				if isinstance(datum, float) or isinstance(datum, int):
					datum = '{0:.3f}'.format(datum)
				tw_data.setItem(0, i, QTableWidgetItem(datum))
			tw_data.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
			tw_data.resizeColumnToContents(len(headers) - 1)

		try:
			if config.FILE_LOADED:
				# check if previously saved params exist
				default_file_exists = try_default_path()
				if default_file_exists:
					# create an interface for choosing saved parameters
					win = QDialog()
					_layout = QGridLayout()

					l = QComboBox()
					tw_data = QTableWidget()

					headers = []
					for header in self.worksheet['B1':get_column_letter(self.worksheet.max_column)+'1']:
						for item in header:
							headers.append(item.value)

					# swap comment & first item in the headers
					a, b = headers.index('Comments'), headers.index('condition')
					headers[a], headers[b] = headers[b], headers[a]

					# set table widget properties
					tw_data.setEditTriggers(QAbstractItemView.NoEditTriggers)
					tw_data.setColumnCount(len(headers))
					tw_data.setHorizontalHeaderLabels(headers)
					tw_data.verticalHeader().hide()
					tw_data.insertRow(0)

					for cell in self.worksheet.iter_rows(min_col=1, max_col=2, min_row=2, max_row=self.worksheet.max_row):
						l.addItem("{0} - {1}".format(cell[0].value.strftime('%Y-%m-%d %H:%M:%S'), cell[1].value))
					data = []
					for cell in self.worksheet.iter_rows(min_col=2, max_col=self.worksheet.max_column, min_row=2, max_row=self.worksheet.max_row):
						tmp = []
						for item in cell:
							tmp.append(item.value)
						tmp[-1], tmp[0] = tmp[0], tmp[-1]
						data.append(tmp)
					# set default combo box index
					l.setCurrentIndex(0)
					# connect data handler
					l.currentIndexChanged.connect(handle_data_list)

					# set table widget contents to default combo box index
					resizer = tw_data.horizontalHeader()
					for i, datum in enumerate(data[0]):
						if isinstance(datum, float) or isinstance(datum, int):
							datum = '{0:.2f}'.format(datum)
						tw_data.setItem(0, i, QTableWidgetItem(datum))
						resizer.setSectionResizeMode(i, QHeaderView.ResizeToContents)

					# resize tabled widget columns to contents
					tw_data.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
					tw_data.resizeColumnToContents(len(headers) - 1)

					_layout.addWidget(l, 0, 0)
					_layout.addWidget(tw_data, 0, 1, 2, 1)
					win.setLayout(_layout)

					button = QDialogButtonBox(
						QDialogButtonBox.Ok | QDialogButtonBox.Cancel,
						Qt.Horizontal, win
					)
					button.accepted.connect(accept)
					button.rejected.connect(reject)
					_layout.addWidget(button, 1, 0)

					win.setAttribute(Qt.WA_DeleteOnClose)
					win.exec_()
				else:
					# TODO : Implement alternative method to import saved params
					raise NotImplementedError("Predesigned excel file not found at default location to load parameters. Non-default file handler not implemented (yet).")
					# default file not exists - let user choose file?
					# options = QFileDialog.Options()
					# options |= QFileDialog.DontUseNativeDialog
					# files, _ = QFileDialog.getOpenFileNames(
					# 	QFileDialog(),
					# 	"Load Parameters",
					# 	"",
					# 	"Excel Files (*.xlsx);;All Files (*)",
					# 	options=options
					# )

					# do magic here
					# if validate_file():
					# 	pass
					# else:
					# 	raise CustomError("Invalid excel file", loc="paramLoad")
		except CustomError as ce:
			now = datetime.now().replace(microsecond=0)
			print("[{0}] {1}".format(now, ce.message))
		except InvalidFileException as ife:
			print("[INVALID FILE ERROR] {0}".format(ife))
		except SheetTitleException as ste:
			print("[SHEET TITLE ERROR] {0}".format(ste))
		except NotImplementedError as nie:
			print("[NOT IMPLEMENTED ERROR] {0}".format(nie))
		except Exception as e:
			print("[DEBUG MESSAGE] {0}".format(e))
