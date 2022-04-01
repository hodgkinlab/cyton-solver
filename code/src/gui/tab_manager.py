import os
import openpyxl
import numpy as np
from datetime import datetime
from openpyxl.utils import get_column_letter
from collections import deque
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import QWidget, QTabWidget, QHBoxLayout, QGridLayout, QCheckBox, QDoubleSpinBox, QScrollArea, QPlainTextEdit, QPushButton, QDialog, QTreeWidget, QTreeWidgetItem, QDialogButtonBox, QTableWidget, QAbstractItemView, QTableWidgetItem, QAbstractScrollArea, QHeaderView, QLabel

# my modules
import src.common.settings as config
from src.qmods.dslider import QDoubleSlider as DoubleSlider
from src.plot.c1_canvas import C1Canvas
from src.plot.c15_canvas import C15Canvas
from src.plot.cc_canvas import CompareCanvas
from src.gui.button_manager import ButtonManager
from src.gui.event_handlers import C1EventHandler, C15EventHandler
from src.gui.tab_items.c1_items import create_undivided_layout, create_dividing_layout, create_misc_layout
from src.gui.tab_items.c15_items import create_unstimulated_layout, create_stimulated_layout, create_ext_layout
from src.IO.error_handler import CustomError


class TabManager:
	def __init__(self):
		# initialise main tab object to host
		self.main_tabs = QTabWidget()
		self.main_tabs.setTabPosition(QTabWidget.East)
		self.main_tabs.setMovable(True)
		self.main_tabs.setStyleSheet(
			'QTabBar::tab:hover{background-color: #990000;} QTabBar::tab::selected{font: bold;}'
		)

		# self.intro_tab = QWidget()
		self.cyton1_tab = QWidget()
		self.cyton15_tab = QWidget()
		# self.cyton2_tab = QWidget()
		self.compare_tab = QWidget()

		# create event handler objects
		self.c1_event_handler = C1EventHandler()
		self.c15_event_handler = C15EventHandler()

		# create plot canvas objects
		self.c1_plot = C1Canvas()
		self.c1_event_handler.param_changed.connect(self.c1_plot.update_plot)
		self.c15_plot = C15Canvas()
		self.c15_event_handler.param_changed15.connect(self.c15_plot.update_plot)
		self.compare_plot = CompareCanvas()

		# add tabs
		# self.main_tabs.addTab(self.intro_tab, "About")
		self.main_tabs.addTab(self.cyton1_tab, "Cyton 1")
		self.main_tabs.addTab(self.cyton15_tab, "Cyton 1.5")
		# self.main_tabs.addTab(self.cyton2_tab, "Cyton 2")
		self.main_tabs.addTab(self.compare_tab, "Compare")

		# self.intro_tab.setLayout(self.create_intro_tab())
		self.cyton1_tab.setLayout(self.create_cyton1_tab())
		self.cyton15_tab.setLayout(self.create_cyton15_tab())
		# self.cyton2_tab.setLayout(self.create_cyton2_tab())
		self.compare_tab.setLayout(self.create_compare_tab())

		# set default tab to Cyton 1.5 model -> useful for intro tab later on
		self.main_tabs.setCurrentIndex(1)

	# TODO: create a welcome/introduction page of Cyton Solver
	# def create_intro_tab(self):
	# 	PLACE_HOLDER = QPlainTextEdit()
	# 	PLACE_HOLDER.setPlaceholderText(
	# 		"This is a place holder area for Cyton Solver manual. I could put full details of the program if I want! Let's make some graphical 2018 style interactive guide!"
	# 	)
	#
	# 	PLACE_HOLDER.setVisible(True)
	# 	layout = QGridLayout()
	#
	# 	layout.addWidget(PLACE_HOLDER)
	#
	# 	return layout

	def create_cyton1_tab(self):
		"""
		this function will populate cyton1 tab

		define all parameter names
		"""
		self.mu0div = QDoubleSpinBox()
		self.mu0divLock = QCheckBox("lock")
		self.mu0divSlider = DoubleSlider(Qt.Horizontal)
		self.mu0divUpperBound = QDoubleSpinBox()
		self.mu0div.setKeyboardTracking(False)
		self.mu0divUpperBound.setKeyboardTracking(False)

		self.sig0div = QDoubleSpinBox()
		self.sig0divLock = QCheckBox("lock")
		self.sig0divSlider = DoubleSlider(Qt.Horizontal)
		self.sig0divUpperBound = QDoubleSpinBox()
		self.sig0div.setKeyboardTracking(False)
		self.sig0divUpperBound.setKeyboardTracking(False)

		self.mu0death = QDoubleSpinBox()
		self.mu0deathLock = QCheckBox("lock")
		self.mu0deathSlider = DoubleSlider(Qt.Horizontal)
		self.mu0deathUpperBound = QDoubleSpinBox()
		self.mu0death.setKeyboardTracking(False)
		self.mu0deathUpperBound.setKeyboardTracking(False)

		self.sig0death = QDoubleSpinBox()
		self.sig0deathLock = QCheckBox("lock")
		self.sig0deathSlider = DoubleSlider(Qt.Horizontal)
		self.sig0deathUpperBound = QDoubleSpinBox()
		self.sig0death.setKeyboardTracking(False)
		self.sig0deathUpperBound.setKeyboardTracking(False)

		self.muSubdiv = QDoubleSpinBox()
		self.muSubdivLock = QCheckBox("lock")
		self.muSubdivSlider = DoubleSlider(Qt.Horizontal)
		self.muSubdivUpperBound = QDoubleSpinBox()
		self.muSubdiv.setKeyboardTracking(False)
		self.muSubdivUpperBound.setKeyboardTracking(False)

		self.sigSubdiv = QDoubleSpinBox()
		self.sigSubdivLock = QCheckBox("lock")
		self.sigSubdivSlider = DoubleSlider(Qt.Horizontal)
		self.sigSubdivUpperBound = QDoubleSpinBox()
		self.sigSubdiv.setKeyboardTracking(False)
		self.sigSubdivUpperBound.setKeyboardTracking(False)

		self.muSubdeath = QDoubleSpinBox()
		self.muSubdeathLock = QCheckBox("lock")
		self.muSubdeathSlider = DoubleSlider(Qt.Horizontal)
		self.muSubdeathUpperBound = QDoubleSpinBox()
		self.muSubdeath.setKeyboardTracking(False)
		self.muSubdeathUpperBound.setKeyboardTracking(False)

		self.sigSubdeath = QDoubleSpinBox()
		self.sigSubdeathLock = QCheckBox("lock")
		self.sigSubdeathSlider = DoubleSlider(Qt.Horizontal)
		self.sigSubdeathUpperBound = QDoubleSpinBox()
		self.sigSubdeath.setKeyboardTracking(False)
		self.sigSubdeathUpperBound.setKeyboardTracking(False)

		self.pf0 = QDoubleSpinBox()
		self.pf0Lock = QCheckBox("lock")
		self.pf0Slider = DoubleSlider(Qt.Horizontal)
		self.pf0UpperBound = QDoubleSpinBox()
		self.pf0.setKeyboardTracking(False)
		self.pf0UpperBound.setKeyboardTracking(False)

		self.pfmu = QDoubleSpinBox()
		self.pfmuLock = QCheckBox("lock")
		self.pfmuSlider = DoubleSlider(Qt.Horizontal)
		self.pfmuUpperBound = QDoubleSpinBox()
		self.pfmu.setKeyboardTracking(False)
		self.pfmuUpperBound.setKeyboardTracking(False)

		self.pfsig = QDoubleSpinBox()
		self.pfsigLock = QCheckBox("lock")
		self.pfsigSlider = DoubleSlider(Qt.Horizontal)
		self.pfsigUpperBound = QDoubleSpinBox()
		self.pfsig.setKeyboardTracking(False)
		self.pfsigUpperBound.setKeyboardTracking(False)

		self.MechProp = QDoubleSpinBox()
		self.MechPropLock = QCheckBox("lock")
		self.MechPropSlider = DoubleSlider(Qt.Horizontal)
		self.MechPropUpperBound = QDoubleSpinBox()
		self.MechProp.setKeyboardTracking(False)
		self.MechPropUpperBound.setKeyboardTracking(False)

		self.MechDecayConst = QDoubleSpinBox()
		self.MechDecayConstLock = QCheckBox("lock")
		self.MechDecayConstSlider = DoubleSlider(Qt.Horizontal)
		self.MechDecayConstUpperBound = QDoubleSpinBox()
		self.MechDecayConst.setKeyboardTracking(False)
		self.MechDecayConstUpperBound.setKeyboardTracking(False)

		_master_layout = QHBoxLayout()
		_sub_master_layout = QGridLayout()

		_c1_param_layout = QGridLayout()
		_c1_param_layout.addWidget(create_undivided_layout(self), 0, 0, 1, 2)
		_c1_param_layout.addWidget(create_dividing_layout(self), 1, 0, 1, 2)
		_c1_param_layout.addWidget(create_misc_layout(self), 2, 0, 1, 2)

		self.params = [
			self.mu0div, self.sig0div,
			self.mu0death, self.sig0death,
			self.muSubdiv, self.sigSubdiv,
			self.muSubdeath, self.sigSubdeath,
			self.pf0, self.pfmu, self.pfsig,
			self.MechProp, self.MechDecayConst
		]
		self.cyton1_buttons = ButtonManager(parent=self, model_id='cyton1')
		_sub_master_layout.addLayout(self.cyton1_buttons._button_layout, 0, 0, 1, 1)
		_sub_master_layout.addLayout(_c1_param_layout, 1, 0, 10, 1)

		_plot_layout = QGridLayout()

		c1_cellgen_sa = QScrollArea()
		c1_cellgen_sa.setWidgetResizable(True)
		c1_cellgen_sa.setWidget(self.c1_plot.c1_cellgen)

		_plot_layout.addWidget(c1_cellgen_sa, 0, 0, 1, 2)
		_plot_layout.addWidget(self.c1_plot.c1_pdfs, 1, 0)
		_plot_layout.addWidget(self.c1_plot.c1_livecell, 1, 1)

		_master_layout.addLayout(_sub_master_layout)
		_master_layout.addLayout(_plot_layout)

		return _master_layout

	def create_cyton15_tab(self):
		self.unstimMuDeath = QDoubleSpinBox()
		self.unstimMuDeathLock = QCheckBox("lock")
		self.unstimMuDeathSlider = DoubleSlider(Qt.Horizontal)
		self.unstimMuDeathUpperBound = QDoubleSpinBox()
		self.unstimMuDeath.setKeyboardTracking(False)
		self.unstimMuDeathUpperBound.setKeyboardTracking(False)

		self.unstimSigDeath = QDoubleSpinBox()
		self.unstimSigDeathLock = QCheckBox("lock")
		self.unstimSigDeathSlider = DoubleSlider(Qt.Horizontal)
		self.unstimSigDeathUpperBound = QDoubleSpinBox()
		self.unstimSigDeath.setKeyboardTracking(False)
		self.unstimSigDeathUpperBound.setKeyboardTracking(False)

		self.stimMuDiv = QDoubleSpinBox()
		self.stimMuDivLock = QCheckBox("lock")
		self.stimMuDivSlider = DoubleSlider(Qt.Horizontal)
		self.stimMuDivUpperBound = QDoubleSpinBox()
		self.stimMuDiv.setKeyboardTracking(False)
		self.stimMuDivUpperBound.setKeyboardTracking(False)

		self.stimSigDiv = QDoubleSpinBox()
		self.stimSigDivLock = QCheckBox("lock")
		self.stimSigDivSlider = DoubleSlider(Qt.Horizontal)
		self.stimSigDivUpperBound = QDoubleSpinBox()
		self.stimSigDiv.setKeyboardTracking(False)
		self.stimSigDivUpperBound.setKeyboardTracking(False)

		self.stimMuDeath = QDoubleSpinBox()
		self.stimMuDeathLock = QCheckBox("lock")
		self.stimMuDeathSlider = DoubleSlider(Qt.Horizontal)
		self.stimMuDeathUpperBound = QDoubleSpinBox()
		self.stimMuDeath.setKeyboardTracking(False)
		self.stimMuDeathUpperBound.setKeyboardTracking(False)

		self.stimSigDeath = QDoubleSpinBox()
		self.stimSigDeathLock = QCheckBox("lock")
		self.stimSigDeathSlider = DoubleSlider(Qt.Horizontal)
		self.stimSigDeathUpperBound = QDoubleSpinBox()
		self.stimSigDeath.setKeyboardTracking(False)
		self.stimSigDeathUpperBound.setKeyboardTracking(False)

		self.stimMuDD = QDoubleSpinBox()
		self.stimMuDDLock = QCheckBox("lock")
		self.stimMuDDSlider = DoubleSlider(Qt.Horizontal)
		self.stimMuDDUpperBound = QDoubleSpinBox()
		self.stimMuDD.setKeyboardTracking(False)
		self.stimMuDDUpperBound.setKeyboardTracking(False)

		self.stimSigDD = QDoubleSpinBox()
		self.stimSigDDLock = QCheckBox("lock")
		self.stimSigDDSlider = DoubleSlider(Qt.Horizontal)
		self.stimSigDDUpperBound = QDoubleSpinBox()
		self.stimSigDD.setKeyboardTracking(False)
		self.stimSigDDUpperBound.setKeyboardTracking(False)

		self.SubDivTime = QDoubleSpinBox()
		self.SubDivTimeLock = QCheckBox("lock")
		self.SubDivTimeSlider = DoubleSlider(Qt.Horizontal)
		self.SubDivTimeUpperBound = QDoubleSpinBox()
		self.SubDivTime.setKeyboardTracking(False)
		self.SubDivTimeUpperBound.setKeyboardTracking(False)

		self.pfrac = QDoubleSpinBox()
		self.pfracLock = QCheckBox("lock")
		self.pfracSlider = DoubleSlider(Qt.Horizontal)
		self.pfracUpperBound = QDoubleSpinBox()
		self.pfrac.setKeyboardTracking(False)
		self.pfracUpperBound.setKeyboardTracking(False)

		_master_layout = QGridLayout()
		_sub_master_layout = QGridLayout()

		_c15_param_layout = QGridLayout()
		_c15_param_layout.addWidget(create_unstimulated_layout(self), 0, 0, 1, 2)
		_c15_param_layout.addWidget(create_stimulated_layout(self), 1, 0, 3, 2)
		_c15_param_layout.addWidget(create_ext_layout(self), 4, 0, 1, 2)

		self.params15 = [
			self.unstimMuDeath, self.unstimSigDeath,
			self.stimMuDiv, self.stimSigDiv,
			self.stimMuDeath, self.stimSigDeath,
			self.stimMuDD, self.stimSigDD,
			self.SubDivTime, self.pfrac
		]
		self.cyton15_buttons = ButtonManager(parent=self, model_id='cyton1.5')
		_sub_master_layout.addLayout(self.cyton15_buttons._button_layout, 0, 0, 1, 1)
		_sub_master_layout.addLayout(_c15_param_layout, 1, 0, 10, 1)

		_plot_layout = QGridLayout()

		c15_cellgen_sa = QScrollArea()
		c15_cellgen_sa.setWidgetResizable(True)
		c15_cellgen_sa.setWidget(self.c15_plot.c15_cellgen)

		_plot_layout.addWidget(c15_cellgen_sa, 0, 0, 1, 2)
		_plot_layout.addWidget(self.c15_plot.c15_pdfs, 1, 0)
		_plot_layout.addWidget(self.c15_plot.c15_livecell, 1, 1)

		_master_layout.addLayout(_sub_master_layout, 0, 0)
		_master_layout.addLayout(_plot_layout, 0, 1)

		# TODO: perhaps convert parameter control section to dockwidget - this requires to reconcile with Cyton 1
		# example_layout = QHBoxLayout()
		# example_layout.addLayout(_sub_master_layout)
		# # example_layout.addLayout(_plot_layout)
		# a = QWidget()
		# self.example_dock = QDockWidget("Example", self.cyton15_tab)
		# self.example_dock.setAllowedAreas(Qt.LeftDockWidgetArea)
		# self.example_dock.setWidget(a)
		# self.example_dock.setFloating(False)
		# self.example_dock.setFeatures(QDockWidget.DockWidgetFloatable | QDockWidget.DockWidgetMovable)
		# a.setLayout(example_layout)
		# a.setSizePolicy(QSizePolicy.MinimumExpanding, QSizePolicy.MinimumExpanding)
		# _master_layout.addWidget(self.example_dock, 0, 0)
		# # _master_layout.addWidget(a, 0, 0)
		#
		# _master_layout.addLayout(_plot_layout, 0, 1)

		return _master_layout

	# TODO: complete Cyton 2 model!
	# def create_cyton2_tab(self):
	#     pass

	def create_compare_tab(self):
		"""
		This function constructs compare tab as well as the tab specific event handlers.

		TODO: split creation and event handling functions apart for better readability.
		:return: (QGridLayout) PyQt5 layout object
		"""
		def handle_select_data():
			def accept():
				# collect selected data
				selected_data = []
				for s, t in enumerate(t_items):
					if t.checkState(0) == Qt.Checked:
						selected_data.append(data[s])

				# TODO: send update signal if and only if 1 or more data are selected
				self.compare_plot.update(selected_data)
				win.close()

			def reject():
				win.close()

			try:
				# prevent using this feature if no data is imported
				if not config.FILE_LOADED:
					raise CustomError("No data to compare! Please import data first.", loc='compare_data')
				else:
					# find the default excel file generated from "Save Params" button
					file_name, file_ext = os.path.splitext(config.FULL_FILE_PATH)
					f = file_name + '_params' + file_ext

					if os.path.isfile(f):
						wb = openpyxl.load_workbook(f)
						# TODO: for now it's only designed for Cyton 1.5. Include Cyton 1 later
						c15_ws = wb['Cyton1.5 Fit']

						win = QDialog()
						_layout = QGridLayout()

						"""Tree Widget for selecting data"""
						tree = QTreeWidget()
						tree.setSelectionMode(QAbstractItemView.SingleSelection)
						tree.setHeaderHidden(True)
						"""Table widget for visualising available data"""
						tw_data = QTableWidget()
						headers = deque([])
						for header in c15_ws['B1':get_column_letter(c15_ws.max_column)+'1']:
							for item in header:
								headers.append(item.value)

						# swap comment & first item in the headers
						# a, b = headers.index('Comments'), headers.index('condition')
						# headers[a], headers[b] = headers[b], headers[a]
						headers.rotate(1)
						headers[0], headers[1] = headers[1], headers[0]

						# set table widget properties
						tw_data.setEditTriggers(QAbstractItemView.NoEditTriggers)
						tw_data.setColumnCount(len(headers))
						tw_data.setHorizontalHeaderLabels(headers)

						data, t_items = [], []
						for cell in c15_ws.iter_rows(min_col=2, max_col=c15_ws.max_column, min_row=2, max_row=c15_ws.max_row):
							tmp = deque([])
							for item in cell:
								tmp.append(item.value)
							tmp.rotate(1)
							tmp[0], tmp[1] = tmp[1], tmp[0]
							data.append(tmp)

						resizer = tw_data.horizontalHeader()
						for i, l in enumerate(data):
							tw_data.insertRow(i)
							tree_item = QTreeWidgetItem(tree)
							tree_item.setText(0, "{0}".format(i+1))
							tree_item.setFlags(tree_item.flags() | Qt.ItemIsUserCheckable)
							tree_item.setCheckState(0, Qt.Unchecked)
							t_items.append(tree_item)

							for j, datum in enumerate(l):
								if isinstance(datum, float) or isinstance(datum, int):
									datum = '{0:.2f}'.format(datum)
								tw_data.setItem(i, j, QTableWidgetItem(datum))
								resizer.setSectionResizeMode(j, QHeaderView.ResizeToContents)

						# resize tabled widget columns to contents
						tw_data.setSizeAdjustPolicy(QAbstractScrollArea.AdjustToContents)
						tw_data.setSizeAdjustPolicy(QAbstractItemView.AdjustToContents)
						tw_data.resizeColumnToContents(len(headers) - 1)

						_layout.addWidget(tree, 0, 0)
						_layout.addWidget(tw_data, 0, 1, 1, 1)

						win.setLayout(_layout)

						button = QDialogButtonBox(
							QDialogButtonBox.Ok | QDialogButtonBox.Cancel,
							Qt.Horizontal, win
						)
						button.accepted.connect(accept)
						button.rejected.connect(reject)
						_layout.addWidget(button, 1, 1)

						win.setAttribute(Qt.WA_DeleteOnClose)
						win.exec_()
					else:
						raise CustomError("There is no saved parameters! This module uses default excel file generated from saving parameters in one of Cyton tabs.", loc='compare_data_no_file')
			except CustomError as ce:
				now = datetime.now().replace(microsecond=0)
				print("[{0}] {1}".format(now, ce))
			except Exception as e:
				print("[ERROR] {0}".format(e))

		def handle_plot_setting():
			def accept():
				self.compare_plot.apply_plot_settings(vals.value())
				w.close()

			def reject():
				w.close()

			w = QDialog()
			_setting_layout = QGridLayout()

			vals = QDoubleSpinBox()
			vals.setMinimumWidth(100)
			vals.setRange(0.0, np.inf)
			vals.setValue(150.00)
			vals.setSingleStep(0.01)
			vals.setSuffix(' hrs')

			set_buttons = QDialogButtonBox(
				QDialogButtonBox.Ok | QDialogButtonBox.Cancel,
				Qt.Horizontal, w
			)
			set_buttons.accepted.connect(accept)
			set_buttons.rejected.connect(reject)

			_setting_layout.addWidget(QLabel('End time :'), 0, 0)
			_setting_layout.addWidget(vals, 0, 1)
			_setting_layout.addWidget(set_buttons, 1, 0)
			w.setLayout(_setting_layout)

			w.setAttribute(Qt.WA_DeleteOnClose)
			w.exec_()

		def handle_clear():
			self.compare_plot.clear_all()

		sub_tab = QTabWidget()
		_layout = QGridLayout()

		select = QPushButton("Select Data")
		select.clicked.connect(handle_select_data)
		setting = QPushButton("Plot Setting")
		setting.clicked.connect(handle_plot_setting)
		clear = QPushButton("Clear All")
		clear.clicked.connect(handle_clear)

		_layout.addWidget(select, 0, 0)
		_layout.addWidget(setting, 0, 1)
		_layout.addWidget(clear, 0, 2)

		# PDF comparison - sub tab 1
		compare_pdfs = QWidget()
		_pdf_layout = QGridLayout()
		_pdf_layout.addWidget(self.compare_plot.pdf1, 0, 0)
		_pdf_layout.addWidget(self.compare_plot.pdf2, 0, 1)
		_pdf_layout.addWidget(self.compare_plot.pdf3, 1, 0)
		_pdf_layout.addWidget(self.compare_plot.pdf4, 1, 1)
		compare_pdfs.setLayout(_pdf_layout)

		# Dose response curves of parameters - sub tab 2
		dose_response = QWidget()

		sub_tab.addTab(compare_pdfs, 'PDFs')
		sub_tab.addTab(dose_response, 'Parameters')

		_layout.addWidget(sub_tab, 1, 0, 1, 3)

		return _layout
